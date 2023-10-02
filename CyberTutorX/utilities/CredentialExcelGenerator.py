from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, Protection, PatternFill


class CredentialExcelGenerator:
    """
        student_info = [{'Full Name': "Mohammad Asad", 'Registration Number': '12123', 'Session': "21-22", 'Class':12,
                            'Section':"A", 'Username':202212121, 'Password':21212121},
                   {'Full Name': "Mohammad Asad", 'Registration Number': '12123', 'Session': "21-22", 'Class':12,
                            'Section':"A", 'Username':202212121, 'Password':21212121},
                   {'Full Name': "Mohammad Asad", 'Registration Number': '12123', 'Session': "21-22", 'Class':12,
                            'Section':"A", 'Username':202212121, 'Password':21212121}
                   ]

        file_name = Credentials.xlsx
    """

    def __init__(self, file_name="Credentials.xlsx"):
        self.__file_name = file_name
        self.__column_width = 20
        self.__row_height = 20
        self.__workbook = openpyxl.Workbook()

    def __create_credential_sheet_excel(self, student_info):
        details_sheet_title = "Credentials"
        active_sheet = self.__workbook.create_sheet(details_sheet_title, 0)
        active_sheet.protection.sheet = True

        header = list(tuple(student_info[0].keys()))
        rows = [tuple(entry.values()) for entry in student_info]

        active_sheet.append(header)

        for row in rows:
            active_sheet.append(row)

        for column in active_sheet.iter_cols():
            active_sheet.column_dimensions[column[0].column_letter].width = self.__column_width
            for cell in column:
                if cell.value:
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='D3D3D3',
                                            end_color='D3D3D3')
                else:
                    cell.protection = Protection(locked=False)

        # changing style of heading
        for cell in active_sheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type='solid',
                                    start_color='BFBFBF',
                                    end_color='BFBFBF')

        for i in range(1, len(rows) + 1):
            active_sheet.row_dimensions[i].height = self.__row_height

    def generate_excel(self, student_info):
        try:
            self.__create_credential_sheet_excel(student_info)
            virtual_workbook = BytesIO()
            self.__workbook.save(virtual_workbook)
            return virtual_workbook
            # self.__workbook.save(self.__file_name)
        except Exception as e:
            print(repr(e))
            # TODO log error f"Exception occured in excel generator {repr(e)}"
            return
