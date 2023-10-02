import openpyxl
from openpyxl.styles import Font, Alignment, Protection, PatternFill
from io import BytesIO


class ResultExcelGenerator:
    """

    detail_values = {
    "Class": 1,
    "Class Teacher": None,
    "Created By": None,
    "Max Days of Attendance": None,
    "Section": "A",
    "Session": "22-23",
    "Test": "Unit 1"
    }

    subject_teacher_marks_value = [{'Subjects': 'English', 'Teacher': 'Jyoti Rani', 'Subject Type': 'Academics'},
                               {'Subjects': 'EVS', 'Teacher': 'Jyoti Rani', 'Subject Type': 'Academics'},
                               {'Subjects': 'Hindi', 'Teacher': 'Jyoti Rani', 'Subject Type': 'Academics'},
                               {'Subjects': 'Mathematics', 'Teacher': 'Jyoti Rani', 'Subject Type': 'Academics'}
                               ]

    student_info = [{'Student_id': 1864, 'Student Name': 'Sambhavya Sharma', 'Roll Number': 1123},
                   {'Student_id': 1865, 'Student Name': 'Sanskar Singh', 'Roll Number': 1125},
                   {'Student_id': 1866, 'Student Name': 'Anay Mahipal', 'Roll Number': 1108},
                   ]
    academic_subject= ["Subject1","Subject2", "Subject3"]
    co_scholastic_subject = ["Subject1","Subject2", "Subject3"]
    co_curricular_subject = ["Subject1","Subject2", "Subject3"]

    """

    def __init__(self, file_name: str, detail_values: dict, subject_teacher_marks_value: list[dict], student_info: list[dict]):
        self.__file_name = file_name
        self.__detail_values = detail_values
        self.__subject_teacher_marks_value = subject_teacher_marks_value
        self.__student_info = student_info
        self.__workbook = openpyxl.Workbook()
        self.__column_width = 20
        self.__row_height = 20

    def __create_detail_sheet_excel(self):
        details_sheet_title = "Details"
        active_sheet = self.__workbook.create_sheet(details_sheet_title, 0)
        active_sheet.protection.sheet = True

        # Adding rows in according to detail_value dict
        for row in list(self.__detail_values.items()):
            active_sheet.append(row)

        for column in active_sheet.iter_cols():
            active_sheet.column_dimensions[column[0].column_letter].width = self.__column_width
            for cell in column:
                if cell.value:
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='D3D3D3',
                                            end_color='D3D3D3')
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.protection = Protection(locked=False)

    def __create_subject_teacher_marks_info(self):
        subject_teacher_marks_sheet_title = "Subject Info"
        total_columns = 3
        active_sheet = self.__workbook.create_sheet(subject_teacher_marks_sheet_title, 1)
        active_sheet.protection.sheet = True

        if len(self.__subject_teacher_marks_value) <= 0:
            raise RuntimeError("subject_teacher_marks_value has zero values")
        if len(self.__subject_teacher_marks_value[0]) != total_columns:
            raise RuntimeError(f"subject_teacher_marks_value has other than {total_columns}")

        header = list(tuple(self.__subject_teacher_marks_value[0].keys())) + ["Min Marks", "Max Marks"]
        rows = [tuple(entry.values()) for entry in self.__subject_teacher_marks_value]

        active_sheet.append(header)

        # Adding rows in sheet
        for row in rows:
            active_sheet.append(row)

        # Disabling protection and adding color
        for column in active_sheet.iter_cols():
            active_sheet.column_dimensions[column[0].column_letter].width = self.__column_width
            for cell in column:
                if cell.value:
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='D3D3D3',
                                            end_color='D3D3D3')
                else:
                    cell.protection = Protection(locked=False)

        # changing style of heading
        for cell in active_sheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type='solid',
                                    start_color='BFBFBF',
                                    end_color='BFBFBF')

        # increasing row height
        for i in range(1, len(rows) + 1):
            active_sheet.row_dimensions[i].height = self.__row_height

    def __create_student_marks_info(self, academic_subject: list, co_scholastic_subject: list, co_curricular_subject:list):
        student_marks_sheet_title = "Result"
        total_columns = 2
        active_sheet = self.__workbook.create_sheet(student_marks_sheet_title, 2)
        active_sheet.protection.sheet = True
        if len(self.__student_info) <= 0:
            raise RuntimeError("student_info has zero values")
        if len(self.__student_info[0]) != total_columns:
            raise RuntimeError(f"student_info has other than {total_columns}")
        header = list(tuple(self.__student_info[0].keys())) + ["Attendance"] + academic_subject + \
                 co_scholastic_subject + \
                 co_curricular_subject + ["Remark"]
        rows = [tuple(entry.values()) for entry in self.__student_info]
        active_sheet.append(header)

        # Adding rows in sheet
        for row in rows:
            active_sheet.append(row)

        # Disabling protection and adding color
        for column in active_sheet.iter_cols():
            active_sheet.column_dimensions[column[0].column_letter].width = self.__column_width
            for cell in column:
                if cell.value:
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='D3D3D3',
                                            end_color='D3D3D3')
                else:
                    cell.protection = Protection(locked=False)

        # changing style of heading

        header_color = ["D3D3D3"] * len(tuple(self.__student_info[0].keys())) + ["D3D3D3"] + ["53DCA2"] * len(
            academic_subject) \
                       + ["65D7FF"] * len(co_scholastic_subject) + ["24BECB"] * len(co_curricular_subject) + ["D3D3D3"]

        column_number = 0
        for rows in active_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(header)):
            for cell in rows:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type='solid',
                                        start_color=header_color[column_number],
                                        end_color=header_color[column_number])
                column_number += 1

        # increasing row height
        for i in range(1, len(rows) + 1):
            active_sheet.row_dimensions[i].height = self.__row_height

    def generate_excel(self, academic_subject, co_scholastic_subject, co_curricular_subject):
        try:
            self.__create_detail_sheet_excel()
            self.__create_subject_teacher_marks_info()
            self.__create_student_marks_info(academic_subject, co_scholastic_subject, co_curricular_subject)
            virtual_workbook = BytesIO()
            self.__workbook.save(virtual_workbook)
            return virtual_workbook,True
            # self.__workbook.save(self.__file_name)
        except Exception as e:

            # TODO log error f"Exception occured in excel generator {repr(e)}"
            return str(e),False


