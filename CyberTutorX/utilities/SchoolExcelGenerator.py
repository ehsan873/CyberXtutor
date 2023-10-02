import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO


class SchoolExcelGenerator:
    """
    class_data = ["Class1", "Class2", "Class3"]
    section_data = ["Section1", "Section2", "Section3"]
    session_data = ["Session1","Session2","Session3"]
    teacher_data = ["Teacheremail", "Teacheremail"]

    """

    gender = ["Male", "Female"]
    subject_type = ["Academics", "Co-Curricular", "Co-Scholastic"]
    pre_filled_teacher_data = ["Full Name", "Email", "Phone Number", "Address", "Gender",
                               "Is Class Teacher?",
                               "Class", "Section"]
    pre_filled_student_data = ["First Name", "Middle Name", "Last Name", "Date of Birth", "Email",
                               "Registration Number", "Father Name", "Mother Name", "Phone Number", "Address Line 1",
                               "Address Line 2", "City", "State", "Country", "Pincode",
                               "Gender", "Session", "Class", "Section",
                               ]
    pre_filled_subject_data = ["Subject Name", "Subject Type", "Class", "Section", "Teacher Email", "Session"]
    boolean = ["Yes", "No"]

    def __init__(self):
        self.__column_width = 20

    def __create_teacher_data_sheet_excel(self, class_data: list, section_data: list) -> openpyxl.workbook:
        workbook = openpyxl.Workbook()
        teacher_sheet_title = "Teachers Data"
        active_sheet = workbook.create_sheet(teacher_sheet_title, 0)
        mandatory_heading = ["Mandatory", "Mandatory", "Mandatory", None, "Mandatory",
                             "Mandatory", "If Class Teacher is 'Yes'", "If Class Teacher is 'Yes'"]

        active_sheet.append(mandatory_heading)
        active_sheet.append(self.pre_filled_teacher_data)

        self.__format_headers(active_sheet)

        self.__set_column_width(active_sheet, self.__column_width)

        self.__add_cell_validation(active_sheet, class_data, 'G3:G1048576')
        self.__add_cell_validation(active_sheet, section_data, 'H3:H1048576')
        self.__add_cell_validation(active_sheet, self.gender, 'E3:E1048576')
        self.__add_cell_validation(active_sheet, self.boolean, 'F3:F1048576')

        return workbook

    def __create_student_data_sheet_excel(self, session_data: list, class_data: list,
                                          section_data: list) -> openpyxl.workbook:
        student_sheet_title = "Student Data"
        workbook = openpyxl.Workbook()
        active_sheet = workbook.create_sheet(student_sheet_title, 0)
        mandatory_heading = ["Mandatory", None, None, "Mandatory", None, "Mandatory",
                             None, "Mandatory", "Mandatory", "Mandatory", None, "Mandatory",
                             "Mandatory", "Mandatory", "Mandatory", "Mandatory",
                             "Mandatory", "Mandatory",
                             "Mandatory"]
        active_sheet.append(mandatory_heading)
        active_sheet.append(self.pre_filled_student_data)

        self.__format_headers(active_sheet)
        self.__set_column_width(active_sheet, self.__column_width)

        self.__add_cell_validation(active_sheet, self.gender, 'P3:P1048576')
        self.__add_cell_validation(active_sheet, session_data, 'Q3:Q1048576')
        self.__add_cell_validation(active_sheet, class_data, 'R3:R1048576')
        self.__add_cell_validation(active_sheet, section_data, 'S3:S1048576')

        validator = DataValidation(type="date", allow_blank=True)
        validator.error = 'Your entry is not in the Valid DD-MM-YYYY'
        validator.errorTitle = 'Invalid Entry'
        validator.prompt = 'Date of Birth'
        validator.promptTitle = 'DOB'
        active_sheet.add_data_validation(validator)
        validator.showInputMessage = True
        validator.showErrorMessage = True
        validator.add('D3:D1048576')

        return workbook

    def __create_subject_data_sheet_excel(self, class_data: list, section_data: list,
                                          teacher_data: list, session_data:list) -> openpyxl.workbook:
        subject_sheet_title = "Subject Data"
        workbook = openpyxl.Workbook()
        active_sheet = workbook.create_sheet(subject_sheet_title, 0)
        mandatory_heading = ["Mandatory", "Mandatory", "Mandatory", "Mandatory",
                             "Mandatory", "Mandatory"]
        active_sheet.append(mandatory_heading)
        active_sheet.append(self.pre_filled_subject_data)

        self.__format_headers(active_sheet)
        self.__set_column_width(active_sheet, self.__column_width)

        self.__add_cell_validation(active_sheet, teacher_data, 'E3:E1048576')
        self.__add_cell_validation(active_sheet, class_data, 'C3:C1048576')
        self.__add_cell_validation(active_sheet, section_data, 'D3:D1048576')
        self.__add_cell_validation(active_sheet, self.subject_type, 'B3:B1048576')
        self.__add_cell_validation(active_sheet, session_data, 'F3:F1048576')

        return workbook

    def __add_cell_validation(self, active_sheet, validation_value: list, fields: str) -> None:
        validator = DataValidation(type="list", formula1='"' + ",".join(validation_value) + '"', allow_blank=True)
        validator.error = 'Your entry is not in the Dropdown'
        validator.errorTitle = 'Invalid Entry'
        validator.prompt = 'Please select from the Dropdown'
        validator.promptTitle = 'Selection List'
        active_sheet.add_data_validation(validator)
        validator.showInputMessage = True
        validator.showErrorMessage = True
        validator.add(fields)

    def __format_headers(self, active_sheet):
        for rows in active_sheet.iter_rows():
            for cell in rows:
                if cell.row == 1:
                    cell.font = Font(color="FF0000", bold=True)
                else:
                    cell.font = Font(bold=True)

    def __set_column_width(self, active_sheet, width: int) -> None:
        for column in active_sheet.iter_cols():
            active_sheet.column_dimensions[column[0].column_letter].width = width

    def generate_teacher_excel(self, class_data: list, section_data: list) -> BytesIO:
        file_name = "Teacher.xlsx"
        try:
            workbook = self.__create_teacher_data_sheet_excel(class_data, section_data)
            virtual_workbook = BytesIO()
            workbook.save(virtual_workbook)
            return virtual_workbook
            # workbook.save(file_name)
        except Exception as e:
            # TODO log error f"Exception occured in SchoolExcelGenerator_generate_teacher_excel {repr(e)}"
            return

    def generate_student_excel(self, session_data: list, class_data: list, section_data: list) -> BytesIO:
        file_name = "Student.xlsx"
        try:
            workbook = self.__create_student_data_sheet_excel(session_data, class_data, section_data)
            virtual_workbook = BytesIO()
            workbook.save(virtual_workbook)
            return virtual_workbook
            # workbook.save(file_name)
        except Exception as e:
            print(e)
            # TODO log error f"Exception occured in SchoolExcelGenerator_generate_student_excel {repr(e)}"
            return

    def generate_subject_excel(self, class_data: list, section_data: list, teacher_data: list,session_data:list) -> BytesIO:
        file_name = "Subject.xlsx"
        try:
            workbook = self.__create_subject_data_sheet_excel(class_data, section_data, teacher_data,session_data)
            virtual_workbook = BytesIO()
            workbook.save(virtual_workbook)
            return virtual_workbook
            # workbook.save(file_name)
        except Exception as e:
            print(e)
            # TODO log error f"Exception occured in SchoolExcelGenerator_generate_teacher_excel {repr(e)}"
            return